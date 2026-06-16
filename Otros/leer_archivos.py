import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle

# ============================================
# ARCHIVOS
# ============================================

archivo_csv = "processed.1779746409_TRIPLETAS POR BATCH_msisdn.csv"
archivo_xlsx = "archivo_protegido.xlsx"

# ============================================
# LEER CSV COMO TEXTO
# ============================================

df = pd.read_csv(
    archivo_csv,
    dtype=str,
    keep_default_na=False,
    encoding="latin1"
)

# ============================================
# EXPORTAR A EXCEL
# ============================================

df.to_excel(
    archivo_xlsx,
    index=False,
    engine="openpyxl"
)

# ============================================
# FORMATEAR COLUMNAS COMO TEXTO
# ============================================

wb = load_workbook(archivo_xlsx)
ws = wb.active

# Aplicar formato texto a TODAS las celdas
for row in ws.iter_rows():
    for cell in row:
        cell.number_format = "@"

# Guardar cambios
wb.save(archivo_xlsx)

print(f"\nArchivo generado correctamente:")
print(archivo_xlsx)